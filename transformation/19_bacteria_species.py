from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

filename = "Bacteria_2015-2021.xlsx"
result = "Bacteria_2015-2021_out.xlsx"

code_keywords = ["Lineage", "Subgroup", "group", "clade"]
garbage_keywords = ["metagenome", "uncultured", "bacterium"]

wb = load_workbook(filename=filename)
ws = wb.active
replaced = Font(color="FF0000")
original = Font(color="000000")

# Determine the last filled column
last_column = ws.max_column
output_column = last_column + 1
print(output_column)

# Determine the last row
nr_items = ws.max_row
print(nr_items)  # Number of relevant lines


def get_val(x, y):  # Reads from loaded spreadsheet, cell x,y
    s = get_column_letter(x) + str(y)
    string = ws[s].value
    if not (type(string) == str):
        string = ""  # Numbers can be ignored for our purpose
    return string


def set_val(x, y, val, font=original):  # Write to cell x,y
    # s=get_column_letter(x)+str(y)
    ws.cell(column=x, row=y, value=val)
    ws.cell(column=x, row=y).font = font


def is_garbage(name):  # Searches for garbage keywords in name
    name_list = name.split()
    for key in garbage_keywords:
        if key in name_list:  # If name searches for any letter combination identical
            return True  # If name_list searches for entire words
    return False


def is_all_code(name):  # Searches for code keywords in name
    for key in code_keywords:
        if key in name:
            return True
    return False


def is_code(name):
    if "-" in name:  # There is a minus sign anywhere -> code
        return True
    if any(ele.isnumeric() for ele in name):  # number digits? -> code
        return True
    wordlist = name.split(" ")
    for word in wordlist:  # check  upper case
        if any(ele.isupper() for ele in word[1:]):  # second letter caps -> code
            return True  # We assume it is code
    return False  # Maybe latin?


def is_latin(name):
    if name == "":
        return False  # No text at all?
    if "uncultured" in name:
        return False
    return not is_code(name)  # If it isn't code -> it's latin!


# Index of column Kingdom
kingdom_column_index = None
for column in ws.iter_cols():
    for cell in column:
        if cell.value == "Kingdom":
            kingdom_column_index = cell.column
            break

# Main program
for row in range(2, nr_items + 1):  # Go through all lines
    column = last_column
    pot_name = get_val(column, row)
    if "sp." in pot_name:
        set_val(output_column, row, pot_name, original)
        continue  # "sp." appears in CC -> just leave untouched
    if is_latin(pot_name) and (not is_garbage(pot_name)):
        new_name = pot_name  # +" sp." # maybe just species alone -> untouched
        # set_val(column,row,new_name) # Done if untouched
        set_val(output_column, row, pot_name, original)
        continue
    # Non-trivial cases....
    list = pot_name.split(" ")
    if is_code(list[-1]):
        code_part = list[-1]
    else:
        code_part = ""
    found_latin = False
    while (column >= kingdom_column_index + 1) and (not found_latin):
        column -= 1
        name = get_val(column, row)
        if is_all_code(name):
            code_part = name
            continue
        if not is_garbage(name):
            is_all_latin = True
            name_part = ""
            list = name.split(" ")
            for l in list:
                if is_code(l):  # and code_part=="":
                    if code_part == "":
                        code_part = l
                    else:
                        code_part = code_part + "-" + l
                if is_latin(l):
                    if name_part == "":
                        name_part = l
                    else:
                        name_part = name_part + " " + l
                    found_latin = True
                else:
                    is_all_latin = False
    if is_all_latin:
        new_name = name + " sp. " + code_part
        set_val(output_column, row, new_name, replaced)
        continue

    if found_latin:
        new_name = name_part + " sp. " + code_part
        set_val(output_column, row, new_name, replaced)

wb.active = ws
wb.save(filename=result)

print("done")
