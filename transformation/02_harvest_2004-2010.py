import openpyxl
import os
import re
from openpyxl import load_workbook


# Function VLOOKUP with python
def sverweis(lookup_value, lookup_range, return_column_index):
    for row in lookup_range:
        if row[0].value == lookup_value:
            return row[return_column_index].value


## Surrogate keys, to add the harvest data into the LTFE template
path_keys = "Datenaufbereitung/Ernte"
wb_keys = load_workbook(os.path.join(path_keys, "ID-Matching.xlsx"))
ws_keys = wb_keys.active

# Folder path for iteration
path = "Datenaufbereitung/Ernte/2004-2010"

for file_name in os.listdir(path):
    if file_name.endswith(".xlsx"):
        # Open Excel file
        wb_Ernte = load_workbook(os.path.join(path, file_name))
        ws_Ernte = wb_Ernte.active

        # Check if column A is completely empty, as the scans have different formats
        is_column_a_empty = True
        for cell in ws_Ernte["A"]:
            if cell.value:
                is_column_a_empty = False
                break

        # Remove column A if completely empty
        if is_column_a_empty:
            ws_Ernte.delete_cols(1)

        # Search column B for cells containing numbers and the letter "a"
        for i, cell in enumerate(ws_Ernte["B"], start=1):
            if (
                cell.value
                and isinstance(cell.value, str)
                and "a" in cell.value
                and any(char.isdigit() for char in cell.value)
            ):
                cell_value = cell.value
                cell_row = cell.row

                # Find the first number in the cell
                match = re.search(r"\d+", cell_value)
                if match:
                    # Extract the numeric part from the cell
                    numeric_part = match.group()

                    # Add the numeric part in the cell below
                    for i in range(cell_row + 1, cell_row + 4):
                        cell_vlaue = ws_Ernte.cell(row=i, column=2).value
                        if cell_value is not None:
                            ws_Ernte.cell(row=i, column=2).value = (
                                numeric_part + cell_vlaue.lower()
                            )

                        i += 5

            # If only a number is entered in column B, 'a' must be appended and the rest of the process remains the same
            if cell.value and all(char.isdigit() for char in str(cell.value)):
                cell.value = str(cell.value) + "a"

                cell_value = cell.value
                cell_row = cell.row

                # Find the first number in the cell
                match = re.search(r"\d+", cell_value)
                if match:
                    # Extract the numeric part from the cell
                    numeric_part = match.group()

                    # Add the numeric part in the cell below
                    for i in range(cell_row + 1, cell_row + 4):
                        cell_vlaue = ws_Ernte.cell(row=i, column=2).value
                        if cell_value is not None:
                            ws_Ernte.cell(row=i, column=2).value = (
                                numeric_part + cell_vlaue.lower()
                            )

                        i += 5

        # Add to all sheets in column C an empty column. This is required for the surrogate keys
        ws_Ernte.insert_cols(3)

        # Use function "VLOOKUP", to merge the surrogate keys with the harvest data
        for i, cell in enumerate(ws_Ernte["B"], start=5):
            lookup_value = cell.value
            lookup_range = ws_keys["D1:E241"]
            return_column_index = 1

            # Check if the cell got merged
            if cell.coordinate in ws_Ernte.merged_cells:
                # Continue, if merge is performed
                continue
            else:
                sverweis_value = sverweis(
                    lookup_value, lookup_range, return_column_index
                )
                cell_offset = cell.offset(column=1)
                cell_offset.value = sverweis_value

        # Save the open file
        file_path = os.path.join(path, file_name)
        wb_Ernte.save(file_path)
        print(file_name + ": done")
