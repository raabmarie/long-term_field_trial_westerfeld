import openpyxl
import os
from openpyxl import load_workbook

# OTU Table
wb_OTU_Fungi = load_workbook("OTU_Fungi_2019_Mais.xlsx")
ws_OTU_Fungi = wb_OTU_Fungi.active
print("OTU table open")

# Template to prepare the data
wb_temp_FungiData = load_workbook("Prep_2019_Mais.xlsx")
ws_temp_FungiData_Prep = wb_temp_FungiData["Prep"]
ws_temp_FungiData_Params = wb_temp_FungiData["Params"]
ws_temp_FungiData_OTU_FUNGI = wb_temp_FungiData["OTU_FUNGI"]
print("Prep template open")

# Determine the values needed to merge the data
werte_Fungi = [cell.value for cell in ws_temp_FungiData_Prep["K"]]
del werte_Fungi[0]  # Remove the first entry from the list, as it is a column header
max_Fungi_ID = int(werte_Fungi[-1])  # Last entry in the list

werte_Fungi_Name = [cell.value for cell in ws_temp_FungiData_Prep["A"]]
del werte_Fungi_Name[
    0
]  # Remove the first entry from the list, as it is a column header

werte_Param = [cell.value for cell in ws_temp_FungiData_Params["A"]]
del werte_Param[0]  # Remove the first entry from the list, as it is a column header
max_Param_ID = int(werte_Param[-1]) + 1  # Last entry in the list
print("Preparation of the parameter successfully completed.")

# Merge the values for all parameters
Index_Row_FungiID = 2
last_index = len(werte_Fungi_Name) - 1

for i in range(1, max_Param_ID):
    werte_OTU_table = [cell.value for cell in ws_OTU_Fungi[i + 1]]
    werte_OTU_table_insert = werte_OTU_table[: last_index + 1]

    for index, (wertID, wertName, werte_OTU) in enumerate(
        zip(werte_Fungi, werte_Fungi_Name, werte_OTU_table_insert)
    ):
        zelle_FungiID = ws_temp_FungiData_OTU_FUNGI.cell(
            row=Index_Row_FungiID + index, column=1
        )  # Row 2 + index, column 1 (A)
        zelle_FungiID.value = wertID

        zelle_FungiName = ws_temp_FungiData_OTU_FUNGI.cell(
            row=Index_Row_FungiID + index, column=2
        )  # Row 2 + index, column 2 (B)
        zelle_FungiName.value = wertName

        zelle_OTUWert = ws_temp_FungiData_OTU_FUNGI.cell(
            row=Index_Row_FungiID + index, column=3
        )  # Row 2 + index, column 3 (C)
        zelle_OTUWert.value = werte_OTU

        zelle_Param = ws_temp_FungiData_OTU_FUNGI.cell(
            row=Index_Row_FungiID + index, column=4
        )  # Row 2 + index, column 4 (D)
        zelle_Param.value = i
    Index_Row_FungiID = Index_Row_FungiID + max_Fungi_ID
print("Merging of the two files completed")

# Create a new tab for the final data preparation
final_sheet = wb_temp_FungiData.create_sheet("final")

# Copy the header from tab 1 to tab 2
header_row = next(ws_temp_FungiData_OTU_FUNGI.iter_rows(min_row=1, max_row=1))
for cell in header_row:
    final_sheet[cell.coordinate].value = cell.value

print("New tab was added with header")

# Copy the desired rows to the new tab
for row in ws_temp_FungiData_OTU_FUNGI.iter_rows(
    min_row=2,
    max_row=ws_temp_FungiData_OTU_FUNGI.max_row,
    min_col=1,
    max_col=ws_temp_FungiData_OTU_FUNGI.max_column,
):
    if row[2].value != 0:  # Check the value in column C
        values = [cell.value for cell in row]
        final_sheet.append(values)
print("The desired columns were copied")

# Save file
wb_temp_FungiData.save("OTU_Fungi_2019_Mais_final.xlsx")
print("The workbook was saved")
print("Done")
