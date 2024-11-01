import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

ordner_pfad = "Ernte/Ernte_Qualität/2012-2019/WW1_WW2"

# Iterate through all Excel files
for datei_name in os.listdir(ordner_pfad):
    if datei_name.endswith(".xlsx"):
        datei_pfad = os.path.join(ordner_pfad, datei_name)
        print(datei_name)

        # Load the Excel file
        workbook = load_workbook(filename=datei_pfad)
        sheet_RP = workbook["RP"]

        # Add plot name for a1b3
        sheet_RP["A3"] = "D-a1b3-3"
        sheet_RP["A5"] = "D-a1b3-2"
        sheet_RP["A7"] = "D-a1b3-1"

        sheet_RP["A11"] = "C-a1b3-3"
        sheet_RP["A13"] = "C-a1b3-2"
        sheet_RP["A15"] = "C-a1b3-1"

        sheet_RP["A19"] = "B-a1b3-3"
        sheet_RP["A21"] = "B-a1b3-2"
        sheet_RP["A23"] = "B-a1b3-1"

        sheet_RP["A27"] = "A-a1b3-3"
        sheet_RP["A29"] = "A-a1b3-2"
        sheet_RP["A31"] = "A-a1b3-1"

        # Add column C
        sheet_RP.insert_cols(3)

        # Add plot name for a1b1
        sheet_RP["C3"] = "D-a1b1-3"
        sheet_RP["C5"] = "D-a1b1-2"
        sheet_RP["C7"] = "D-a1b1-1"

        sheet_RP["C11"] = "C-a1b1-3"
        sheet_RP["C13"] = "C-a1b1-2"
        sheet_RP["C15"] = "C-a1b1-1"

        sheet_RP["C19"] = "B-a1b1-3"
        sheet_RP["C21"] = "B-a1b1-2"
        sheet_RP["C23"] = "B-a1b1-1"

        sheet_RP["C27"] = "A-a1b1-3"
        sheet_RP["C29"] = "A-a1b1-2"
        sheet_RP["C31"] = "A-a1b1-1"

        # Add plot name for a2b1
        sheet_RP["L3"] = "D-a2b1-3"
        sheet_RP["L5"] = "D-a2b1-2"
        sheet_RP["L7"] = "D-a2b1-1"

        sheet_RP["L11"] = "C-a2b1-3"
        sheet_RP["L13"] = "C-a2b1-2"
        sheet_RP["L15"] = "C-a2b1-1"

        sheet_RP["L19"] = "B-a2b1-3"
        sheet_RP["L21"] = "B-a2b1-2"
        sheet_RP["L23"] = "B-a2b1-1"

        sheet_RP["L27"] = "A-a2b1-3"
        sheet_RP["L29"] = "A-a2b1-2"
        sheet_RP["L31"] = "A-a2b1-1"

        # Add column N
        sheet_RP.insert_cols(14)

        # Add plot name for a2b3
        sheet_RP["N3"] = "D-a2b3-3"
        sheet_RP["N5"] = "D-a2b3-2"
        sheet_RP["N7"] = "D-a2b3-1"

        sheet_RP["N11"] = "C-a2b3-3"
        sheet_RP["N13"] = "C-a2b3-2"
        sheet_RP["N15"] = "C-a2b3-1"

        sheet_RP["N19"] = "B-a2b3-3"
        sheet_RP["N21"] = "B-a2b3-2"
        sheet_RP["N23"] = "B-a2b3-1"

        sheet_RP["N27"] = "A-a2b3-3"
        sheet_RP["N29"] = "A-a2b3-2"
        sheet_RP["N31"] = "A-a2b3-1"

        # Create a matrix A3-B31
        matrix_a1b3 = []  # 2 x 29 matrix
        for row in range(3, 32):
            row_values = []
            for column in range(1, 3):
                cell_value = sheet_RP.cell(row=row, column=column).value
                row_values.append(cell_value)
            matrix_a1b3.append(row_values)

        # Create a matrix C3-D31
        matrix_a1b1 = []  # 2 x 29 matrix
        for row in range(3, 32):
            row_values = []
            for column in range(3, 5):
                cell_value = sheet_RP.cell(row=row, column=column).value
                row_values.append(cell_value)
            matrix_a1b1.append(row_values)

        # Create a matrix L3-M31
        matrix_a2b1 = []  # 2 x 29 matrix
        for row in range(3, 32):
            row_values = []
            for column in range(12, 14):
                cell_value = sheet_RP.cell(row=row, column=column).value
                row_values.append(cell_value)
            matrix_a2b1.append(row_values)

        # Create a matrix N3-O31
        matrix_a2b3 = []  # 2 x 29 matrix
        for row in range(3, 32):
            row_values = []
            for column in range(14, 16):
                cell_value = sheet_RP.cell(row=row, column=column).value
                row_values.append(cell_value)
            matrix_a2b3.append(row_values)

        # Create a new tab named RP_new
        workbook.create_sheet("RP_new")
        sheet_RP_new = workbook["RP_new"]

        # Insert matrices into the new tab RP_new, column A-B
        for i, row_values in enumerate(matrix_a1b3):  # iterate all 29 rows
            for j, cell_value in enumerate(row_values):  # iterate both columns
                sheet_RP_new.cell(row=i + 1, column=j + 1, value=cell_value)

        for i, row_values in enumerate(matrix_a1b1):  # iterate all 29 rows
            for j, cell_value in enumerate(row_values):  # iterate both columns
                sheet_RP_new.cell(row=i + 30, column=j + 1, value=cell_value)

        for i, row_values in enumerate(matrix_a2b1):  # iterate all 29 rows
            for j, cell_value in enumerate(row_values):  # iterate both columns
                sheet_RP_new.cell(row=i + 59, column=j + 1, value=cell_value)

        for i, row_values in enumerate(matrix_a2b3):  # iterate all 29 rows
            for j, cell_value in enumerate(row_values):  # iterate both columns
                sheet_RP_new.cell(row=i + 88, column=j + 1, value=cell_value)

        # Save file
        workbook.save(datei_pfad)

print("done")
