import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

ordner_pfad = "Ernte/Ernte_Qualität/2012-2019/KM"

# Iterate through all Excel files
for datei_name in os.listdir(ordner_pfad):
    if datei_name.endswith(".xlsx"):
        datei_pfad = os.path.join(ordner_pfad, datei_name)
        print(datei_name)

        # Load Excel file
        workbook = load_workbook(filename=datei_pfad)

        # Iterate through both tabs starch and protein content
        for sheet in workbook.worksheets:
            if sheet.title in ["RP", "Stärke"]:
                # Add plot name for a1b3
                sheet["A3"] = "D-a1b3-3"
                sheet["A5"] = "D-a1b3-2"
                sheet["A7"] = "D-a1b3-1"

                sheet["A11"] = "C-a1b3-3"
                sheet["A13"] = "C-a1b3-2"
                sheet["A15"] = "C-a1b3-1"

                sheet["A19"] = "B-a1b3-3"
                sheet["A21"] = "B-a1b3-2"
                sheet["A23"] = "B-a1b3-1"

                sheet["A27"] = "A-a1b3-3"
                sheet["A29"] = "A-a1b3-2"
                sheet["A31"] = "A-a1b3-1"

                # Add column C
                sheet.insert_cols(3)

                # Add plot name for a1b1
                sheet["C3"] = "D-a1b1-3"
                sheet["C5"] = "D-a1b1-2"
                sheet["C7"] = "D-a1b1-1"

                sheet["C11"] = "C-a1b1-3"
                sheet["C13"] = "C-a1b1-2"
                sheet["C15"] = "C-a1b1-1"

                sheet["C19"] = "B-a1b1-3"
                sheet["C21"] = "B-a1b1-2"
                sheet["C23"] = "B-a1b1-1"

                sheet["C27"] = "A-a1b1-3"
                sheet["C29"] = "A-a1b1-2"
                sheet["C31"] = "A-a1b1-1"

                # Remove entries for b1P, since they are not requested
                for row in sheet.iter_rows(
                    min_row=3, max_row=32, min_col=12, max_col=12
                ):
                    for cell in row:
                        cell.value = None

                # Add plot name for a2b1
                sheet["L3"] = "D-a2b1-3"
                sheet["L5"] = "D-a2b1-2"
                sheet["L7"] = "D-a2b1-1"

                sheet["L11"] = "C-a2b1-3"
                sheet["L13"] = "C-a2b1-2"
                sheet["L15"] = "C-a2b1-1"

                sheet["L19"] = "B-a2b1-3"
                sheet["L21"] = "B-a2b1-2"
                sheet["L23"] = "B-a2b1-1"

                sheet["L27"] = "A-a2b1-3"
                sheet["L29"] = "A-a2b1-2"
                sheet["L31"] = "A-a2b1-1"

                # Add column N
                sheet.insert_cols(14)

                # Add plot name for a2b3
                sheet["N3"] = "D-a2b3-3"
                sheet["N5"] = "D-a2b3-2"
                sheet["N7"] = "D-a2b3-1"

                sheet["N11"] = "C-a2b3-3"
                sheet["N13"] = "C-a2b3-2"
                sheet["N15"] = "C-a2b3-1"

                sheet["N19"] = "B-a2b3-3"
                sheet["N21"] = "B-a2b3-2"
                sheet["N23"] = "B-a2b3-1"

                sheet["N27"] = "A-a2b3-3"
                sheet["N29"] = "A-a2b3-2"
                sheet["N31"] = "A-a2b3-1"

                # Create a matrix A3-B31
                matrix_a1b3 = []  # 2 x 29 matrix
                for row in range(3, 32):
                    row_values = []
                    for column in range(1, 3):
                        cell_value = sheet.cell(row=row, column=column).value
                        row_values.append(cell_value)
                    matrix_a1b3.append(row_values)

                # Create a matrix C3-D31
                matrix_a1b1 = []  # 2 x 29 matrix
                for row in range(3, 32):
                    row_values = []
                    for column in range(3, 5):
                        cell_value = sheet.cell(row=row, column=column).value
                        row_values.append(cell_value)
                    matrix_a1b1.append(row_values)

                # Create a matrix L3-M31
                matrix_a2b1 = []  # 2 x 29 matrix
                for row in range(3, 32):
                    row_values = []
                    for column in range(12, 14):
                        cell_value = sheet.cell(row=row, column=column).value
                        row_values.append(cell_value)
                    matrix_a2b1.append(row_values)

                # Create a matrix N3-O31
                matrix_a2b3 = []  # 2 x 29 matrix
                for row in range(3, 32):
                    row_values = []
                    for column in range(14, 16):
                        cell_value = sheet.cell(row=row, column=column).value
                        row_values.append(cell_value)
                    matrix_a2b3.append(row_values)

                # Create a new tab with suffix _new
                sheet_new = workbook.create_sheet(title=sheet.title + "_new")

                # Insert matrices to column A-B
                for i, row_values in enumerate(matrix_a1b3):  # iterate all 29 rows
                    for j, cell_value in enumerate(row_values):  # iterate both columns
                        sheet_new.cell(row=i + 1, column=j + 1, value=cell_value)

                for i, row_values in enumerate(matrix_a1b1):  # iterate all 29 rows
                    for j, cell_value in enumerate(row_values):  # iterate both columns
                        sheet_new.cell(row=i + 30, column=j + 1, value=cell_value)

                for i, row_values in enumerate(matrix_a2b1):  # iterate all 29 rows
                    for j, cell_value in enumerate(row_values):  # iterate both columns
                        sheet_new.cell(row=i + 59, column=j + 1, value=cell_value)

                for i, row_values in enumerate(matrix_a2b3):  # iterate all 29 rows
                    for j, cell_value in enumerate(row_values):  # iterate both columns
                        sheet_new.cell(row=i + 88, column=j + 1, value=cell_value)

            # Save file
            workbook.save(datei_pfad)

print("done")
