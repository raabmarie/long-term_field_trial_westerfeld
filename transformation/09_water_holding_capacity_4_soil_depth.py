import openpyxl
import os
from openpyxl import load_workbook

path_keys = "Datenstruktur"
wb_keys = load_workbook(os.path.join(path_keys, "wb_keys.xlsx"))
ws_keys = wb_keys.active

# File path
path = "Datenstruktur/Nmin/done/4"

for file_name in os.listdir(path):
    if file_name.endswith(".xlsx"):
        # Open Excel file
        workbook = load_workbook(os.path.join(path, file_name))
        workbook.save("old_" + file_name)
        # Iterate through all visible tabs
        for worksheet in workbook.worksheets:
            if worksheet.title.strip() in ["NFK1", "NFK2", "Nmin Ernte", "Nmin Frühj"]:
                print(file_name, worksheet.title)

                # Remove columns b1P and b3P in each tab
                remove_P_entries = worksheet["21"]
                for cell in remove_P_entries:
                    if cell.value is not None and "P" in cell.value:
                        for row in worksheet.iter_rows(
                            min_row=3,
                            max_row=21,
                            min_col=cell.column,
                            max_col=cell.column,
                        ):
                            for cell in row:
                                cell.value = None

                # Iterate through matrix B3:M20
                cell_range = worksheet["B3:M20"]
                for row in cell_range:
                    for cell in row:
                        cell_value = cell.value

                        # If cell_value is a string
                        if isinstance(cell_value, str):
                            # the content needs to be split, because 4 values were added to one cell
                            numbers = [
                                float(cell.replace(",", "."))
                                for cell in cell_value.split()
                            ]

                            # Add new column in case not all 4 cells to the right are empty
                            if (
                                (
                                    cell.offset(column=1).value is not None
                                    and cell.offset(column=1).value.strip() != ""
                                )
                                or (
                                    cell.offset(column=2).value is not None
                                    and cell.offset(column=2).value.strip() != ""
                                )
                                or (
                                    cell.offset(column=3).value is not None
                                    and cell.offset(column=3).value.strip() != ""
                                )
                                or (
                                    cell.offset(column=4).value is not None
                                    and cell.offset(column=4).value.strip() != ""
                                )
                            ):
                                worksheet.insert_cols(cell.column + 1, len(numbers))

                            # Add values
                            for i, number in enumerate(numbers):
                                worksheet.cell(
                                    row=cell.row, column=cell.column + i
                                ).value = number

                # Copy the surrogate keys fot LTFE template to each tab
                for row in range(1, ws_keys.max_row + 1):
                    valueA = ws_keys.cell(row=row, column=1).value
                    worksheet.cell(row=row, column=1).value = valueA

                    valueC = ws_keys.cell(row=row, column=3).value
                    worksheet.cell(row=row, column=6).value = valueC

                    valueE = ws_keys.cell(row=row, column=5).value
                    worksheet.cell(row=row, column=15).value = valueE

                    valueG = ws_keys.cell(row=row, column=7).value
                    worksheet.cell(row=row, column=20).value = valueG

                # Iterate through columns A, F, O and T and copy all cells including not empty neighbour cells
                copy_values_range = worksheet["2"]
                target_row = 30
                for cell in copy_values_range:
                    if cell.value is not None and "ID" in cell.value:
                        for row in worksheet.iter_rows(
                            min_row=2, min_col=cell.column, max_col=cell.column
                        ):
                            cell = row[0]
                            if cell.value is not None:
                                for i in range(5):
                                    source_cell = worksheet.cell(
                                        row=cell.row, column=cell.column + i
                                    )
                                    target_cell = worksheet.cell(
                                        row=target_row, column=i + 1
                                    )
                                    target_cell.value = source_cell.value
                                target_row = target_row + 1

        # Save file
        workbook.save(file_name)

print("done")
