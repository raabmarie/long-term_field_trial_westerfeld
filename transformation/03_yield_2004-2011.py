import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

ordner_pfad = "Datenaufbereitung/Ernte_Qualtität/2004-2011"

# Iterate through all Excel files in the specified folder
for datei_name in os.listdir(ordner_pfad):
    if datei_name.endswith(".xlsx"):
        datei_pfad = os.path.join(ordner_pfad, datei_name)

        # Load the Excel file
        workbook = load_workbook(filename=datei_pfad)

        # Iterate through all sheets in the Excel file
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Remove table format
            if sheet.tables:
                sheet.tables.clear()

            # Copy C2-D13 to A14-B25
            for row in range(2, 14):
                for col in range(3, 5):
                    dest_cell = sheet.cell(row=row + 12, column=col - 2)
                    src_cell = sheet.cell(row=row, column=col)
                    dest_cell.value = src_cell.value

            # Copy E2-F13 to A26-B37
            for row in range(2, 14):
                for col in range(5, 7):
                    dest_cell = sheet.cell(row=row + 24, column=col - 4)
                    src_cell = sheet.cell(row=row, column=col)
                    dest_cell.value = src_cell.value

            # Copy G2-H13 to A38-B49
            for row in range(2, 14):
                for col in range(7, 9):
                    dest_cell = sheet.cell(row=row + 36, column=col - 6)
                    src_cell = sheet.cell(row=row, column=col)
                    dest_cell.value = src_cell.value

            # Clear columns C-H
            for col in range(3, 9):
                for row in sheet.iter_rows(
                    min_row=1, max_row=sheet.max_row, min_col=col, max_col=col
                ):
                    for cell in row:
                        cell.value = None

            # Set A1 and B1
            sheet["A1"].value = "Parzelle"
            sheet["B1"].value = "Werte"

        # Save the file under the same name
        workbook.save(datei_pfad)

print("done")
