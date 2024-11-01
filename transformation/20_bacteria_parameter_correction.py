import openpyxl

# Open workbooks
# 1. workbook = openpyxl.load_workbook('OTU_Bacteria_2019_total.xlsx')
# 2. workbook = openpyxl.load_workbook('OTU_Bacteria_2020_Mais.xlsx')
# 3. workbook = openpyxl.load_workbook('OTU_Bacteria_2021_Mais.xlsx')
# 4. workbook = openpyxl.load_workbook('OTU_Bacteria_2019_Mais_Gnecco.xlsx')
# 5.
workbook = openpyxl.load_workbook("OTU_Bacteria_2015_WW1_WW2.xlsx")

# Activate tabs "countMatrix" and "final"
count_matrix_sheet = workbook["countMatrix"]
final_sheet = workbook["final"]

# Count number of rows in tab countMatrix
max_rows = count_matrix_sheet.max_row

# 1. Set range B2-DY24283
# 2. Set range B2-BM25986
# 3. Set range B2-BM70235
# 4. Set range B2-AG57838
# 5. Set range B2-AG2801
start_col_idx = 2  # Column B represents index 2
end_col_idx = 33  # Column AG represents index 33

# Start at column B, and iterate through all columns up to and including BM
for col_idx in range(start_col_idx, end_col_idx + 1):
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    column_range = f"{col_letter}2:{col_letter}{max_rows}"

    print(col_letter)
    # Iterate through all rows in the current column
    for row in count_matrix_sheet[column_range]:
        cell_value = row[0].value

        # If the numerical value is not equal to 0, then copy the values to the "final" tab
        if cell_value and cell_value != 0:
            # Copy the values from column A and the current column (col_letter) and the header from columns B, C, ..., BM
            final_sheet.append(
                [
                    count_matrix_sheet[f"A{row[0].row}"].value,
                    count_matrix_sheet[f"{col_letter}1"].value,
                    cell_value,
                ]
            )

# Save file
workbook.save("OTU_Bacteria_2015_WW1_WW2.xlsx")

# Close file
workbook.close()

print("done")
