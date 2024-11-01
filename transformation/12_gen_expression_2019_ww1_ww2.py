import numpy as np
import openpyxl
from openpyxl import load_workbook

# Open Excel file
workbook = load_workbook("2019_Genexpression_WW1_WW2_dct.xlsx")

# Access worksheet dictionary
ws_dct = workbook["dct"]

# Create a new worksheet Trans
ws_Trans = workbook.create_sheet("Trans")

# Step 1: Copy the first row from "dct" to "Trans"
for column_AtoJ in range(1, 11):
    cell_value = ws_dct.cell(row=1, column=column_AtoJ).value
    ws_Trans.cell(row=1, column=column_AtoJ, value=cell_value)

# Step 2: Write the column headers in "Trans"
new_header_values = ["Gene", "Wert", "Kategorie", "Bemerkung"]
for index_header, new_header_value in enumerate(new_header_values, start=11):
    ws_Trans.cell(row=1, column=index_header, value=new_header_value)

# Step 3a): Save the matrix A2-J33 from "dct"
matrix_values_Parzelle = []  # 32 x 10 matrix
for row in range(2, 34):
    row_values = []
    for column in range(1, 11):
        cell_value = ws_dct.cell(row=row, column=column).value
        row_values.append(cell_value)
    matrix_values_Parzelle.append(row_values)

    # Step 3b): Save the values from K1 to AL1 in a list
    header_values_Gene = []  # 1 x 28 list
    for column in range(11, 39):  # column K to AL
        cell_value = ws_dct.cell(row=1, column=column).value
        header_values_Gene.append(cell_value)

# Step 3c): Save the matrix K2-AL33 from "dct"
matrix_values_Werte = np.empty((32, 28))  # 32 x 28 matrix
for row in range(2, 34):
    for column in range(11, 39):  # column K to AL
        cell_value = ws_dct.cell(row=row, column=column).value
        matrix_values_Werte[row - 2, column - 11] = cell_value

# Step 4: For 28 genes, steps 5-7 must be executed
for idx_Gene in range(28):
    start_row = 2 + (
        idx_Gene * 32
    )  # Calculate the starting point for inserting each copy
    end_row = start_row + 31  # Calculate the end point for inserting each copy

    # Step 5: Copy the matrix A2-J33 from "dct" 39 times to "Trans"
    for i, row_values in enumerate(matrix_values_Parzelle):  # iteriere alle 32 Zeilen
        for j, cell_value in enumerate(row_values):  # iteriere alle 10 Spalten
            ws_Trans.cell(row=i + start_row, column=j + 1, value=cell_value)

    # Step 6: Insert the genes in column K named "Gene"
    gene_value = header_values_Gene[idx_Gene]
    for row in range(start_row, end_row + 1):
        ws_Trans.cell(row=row, column=11, value=gene_value)

    # Step 7: Insert the columns of the matrix matrix_values_Werte under each other in column L named "Value"
    start_row_target = start_row
    column_values = matrix_values_Werte[:, idx_Gene]  # Access the column (Index_Gene)

    for value in column_values:
        ws_Trans.cell(row=start_row_target, column=12, value=value)
        start_row_target += 1

# Save file
workbook.save("2019_Genexpression_WW1_WW2_dct.xlsx")
print("done")
