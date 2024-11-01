import openpyxl
import os
import re
from openpyxl import load_workbook


## Function VLOOKUP
def sverweis(lookup_value, lookup_range, return_column_index):
    for row in lookup_range:
        if row[0].value == lookup_value:
            return row[return_column_index].value


## Surrogate keys to add harvest data to the LTFE template
path_keys = "Datenstruktur"
wb_keys = load_workbook(os.path.join(path_keys, "ID-Matching.xlsx"))
ws_keys = wb_keys.active

# folder path
path = "Datenstruktur/ERNTE"

for file_name in os.listdir(path):
    if file_name.endswith(".xlsx"):
        # Open Excel file
        workbook = load_workbook(os.path.join(path, file_name))
        # Save the file with prefix RawData_
        workbook.save("RawData_" + file_name)

        # Iterate through all sheets
        for worksheet in workbook.worksheets:

            # Search in column B for cells with numbers and the letter "a"
            for i, cell in enumerate(worksheet["B"], start=1):
                # Check if the cell has a value, is a string, contains 'a', and contains at least one digit
                if (
                    cell.value
                    and isinstance(cell.value, str)
                    and "a" in cell.value
                    and any(char.isdigit() for char in cell.value)
                ):
                    cell_value = cell.value
                    cell_row = cell.row
                    # Find the first occurrence of digits in the cell
                    match = re.search(r"\d+", cell_value)
                    if match:
                        # Extract the numeric part from the cell
                        numeric_part = match.group()

                        # Insert the numeric part into the cells below
                        for i in range(cell_row + 1, cell_row + 4):
                            cell_vlaue = worksheet.cell(row=i, column=2).value
                            worksheet.cell(row=i, column=2).value = (
                                numeric_part + cell_vlaue
                            )

                    i += 5
                # If only a number is entered in column B, 'a' must be appended and the rest of the process remains the same
                if cell.value and all(char.isdigit() for char in str(cell.value)):
                    cell.value = str(cell.value) + "a"

                    cell_value = cell.value
                    cell_row = cell.row
                    # Find the first occurrence of digits in the cell
                    match = re.search(r"\d+", cell_value)
                    if match:
                        # Extract the numeric part from the cell
                        numeric_part = match.group()

                        # Insert the numeric part into the cells below
                        for i in range(cell_row + 1, cell_row + 4):
                            cell_vlaue = worksheet.cell(row=i, column=2).value
                            worksheet.cell(row=i, column=2).value = (
                                numeric_part + cell_vlaue
                            )

                    i += 5

            # Insert an empty column in column C in each sheet, so that it can be filled with surrogate keys
            worksheet.insert_cols(3)

            # Use function "VLOOKUP" to match the surrogate keys for LTFE template
            for i, cell in enumerate(worksheet["B"], start=5):
                lookup_value = cell.value
                lookup_range = ws_keys["D1:E241"]
                return_column_index = 1
                sverweis_value = sverweis(
                    lookup_value, lookup_range, return_column_index
                )

                cell_offset = cell.offset(column=1)
                cell_offset.value = sverweis_value

        # Save file
        workbook.save(file_name)

print("done")
