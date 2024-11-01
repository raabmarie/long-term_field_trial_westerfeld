from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter

print("1. Definition of Input, Output and Garbage_Keyword")
filename = "041023_Fungi_2015-2021.xlsx"
result = "041023_Fungi_2015-2021_out.xlsx"

garbage_keyword = "unidentified"

replaced = Font(color="FF0000")
original = Font(color="000000")

try:
    wb = load_workbook(filename=filename)
    ws = wb["Fungi"]
except Exception as e:
    print(f"An error occured: {e}")
else:
    print("workbook is open")

print("2. Determine the last column to generate Family_New entries.")
last_column = ws.max_column
output_column = last_column + 1
print(output_column)

print("3. Determine the Family column.")
for column in range(last_column, 0, -1):
    cell = ws.cell(row=1, column=column)
    if cell.value == "Family":
        column_family = ws.cell(row=1, column=column).column
        print(column_family)
        break

print("4. functions")


def get_val(x, y):  # Reads from loaded spreatsheet, cell x,y
    s = get_column_letter(x) + str(y)
    string = ws[s].value
    return string


def set_val(x, y, val, font=original):  # Write to cell x,y
    ws.cell(column=x, row=y, value=val)
    ws.cell(column=x, row=y).font = font


print("5. Number of non empty rows")
nr_items = 1
while bool(get_val(1, nr_items + 1)):
    nr_items += 1  # Counts non-empty lines in spreadsheet
print(nr_items)

print("6. Create a new column with output <> unidentified")
for row in range(2, nr_items + 1):
    column = column_family
    pot_name = get_val(column, row)
    if pot_name != garbage_keyword:
        set_val(output_column, row, pot_name, original)
    else:
        ord_name = get_val(column - 1, row)
        if ord_name != garbage_keyword:
            set_val(output_column, row, ord_name + "_fam", replaced)
        else:
            class_name = get_val(column - 2, row)
            if class_name != garbage_keyword:
                set_val(output_column, row, class_name + "_fam", replaced)
            else:
                phylum_name = get_val(column - 3, row)
                if phylum_name != garbage_keyword:
                    set_val(output_column, row, phylum_name + "_fam", replaced)
                else:
                    set_val(output_column, row, "Fungi_fam", replaced)

wb.save(filename=result)

print("7. done")
